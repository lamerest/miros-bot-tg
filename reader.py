import openpyxl

wb = openpyxl.load_workbook('no brands.xlsx')
sheet = wb.active

wb_brands = openpyxl.load_workbook('all.xlsx')
sheet_brands = wb_brands.active


def len_sort(list):
    for i in range(len(list)):
        for j in range(0, len(list) - 1):
            if len(list[j+1][0]) < len(list[j][0]):
                list[j+1], list[j] = list[j], list[j+1]
    return list


def search(query):
    answer = []
    query = query.replace("ё", "е")

    for row in sheet.iter_rows():
        if query.title() + " " in row[0].value:
            data = []
            for cell in row:
                data.append(str(cell.value).replace("\n", ""))
            answer.append(data)

    for row in sheet.iter_rows():
        if query.title() in row[0].value:
            data = []
            for cell in row:
                data.append(str(cell.value).replace("\n", ""))
            answer.append(data)

    if not answer:
        for row in sheet.iter_rows():
            if query.lower() in row[0].value:
                data = []
                for cell in row:
                    data.append(str(cell.value).replace("\n", ""))
                answer.append(data)

    if not answer:
        long_query = query[0].upper() + query[1:]
        for row in sheet.iter_rows():
            if long_query in row[0].value:
                data = []
                for cell in row:
                    data.append(str(cell.value).replace('\n', ''))
                answer.append(data)

    if not answer and " " in query:
        replaced_query = query.split(' ')[1].title() + " " + query.split(' ')[0].lower()
        for row in sheet.iter_rows():
            if replaced_query in row[0].value:
                data = []
                for cell in row:
                    data.append(str(cell.value).replace('\n', ''))
                answer.append(data)

    return len_sort(answer)


def search_for_brands(query):
    answer = []

    for row in sheet_brands.iter_rows():
        if query.title() + "" in row[0].value:
            data = []
            for cell in row:
                data.append(str(cell.value).replace("\n", ""))
            answer.append(data)

    if not answer:
        for row in sheet_brands.iter_rows():
            if query.replace("ё", "е") in row[0].value:
                data = []
                for cell in row:
                    if '\n' in cell.value:
                        data.append(cell.value.replace('\n', ''))
                    else:
                        data.append(cell.value)
                answer.append(data)

    if not answer:
        for row in sheet.iter_rows():
            if query.lower() in row[0].value:
                data = []
                for cell in row:
                    data.append(str(cell.value).replace("\n", ""))
                answer.append(data)

    if not answer and " " in query:
        long_query = query[0].upper() + query[1:]
        for row in sheet_brands.iter_rows():
            if long_query in row[0].value:
                data = []
                for cell in row:
                    if '\n' in cell.value:
                        data.append(cell.value.replace('\n', ''))
                    else:
                        data.append(cell.value)
                answer.append(data)

    return answer
